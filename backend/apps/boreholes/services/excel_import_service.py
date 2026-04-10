"""钻孔 Excel 导入与数据库查询服务。"""
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from django.conf import settings
from django.contrib.gis.db.models.functions import Distance
from django.contrib.gis.geos import Point
from django.db import transaction
from pyproj import CRS, Transformer

from apps.boreholes.models import Borehole, BoreholeLayer
from apps.boreholes.serializers import BoreholeSerializer

LAYER_FIELD_ALIASES = {
    'borehole_name': ('钻孔名称', '孔号', 'borehole_name', 'name'),
    'layer_name': ('地层名称', '层位', 'layer_name'),
    'depth': ('深度', 'bottom_depth', 'depth'),
    'thickness': ('厚度', 'thickness'),
}

LOCATION_FIELD_ALIASES = {
    'name': ('name', '钻孔名称', '孔号', 'borehole_name'),
    'x': ('x', 'X', '经度', '东坐标', 'lon', 'longitude'),
    'y': ('y', 'Y', '纬度', '北坐标', 'lat', 'latitude'),
    'z': ('z', 'Z', '高程', 'elevation'),
}

LAYER_COLORS = {
    '煤': '#f2c94c',
    '砂': '#56ccf2',
    '泥': '#bb6bd9',
    '土': '#f2994a',
    '风积': '#6fcf97',
}

DATA_DIR = Path(settings.BASE_DIR) / 'data'
BOREHOLE_DIR = DATA_DIR / 'boreholes'
LOCATION_DIR = DATA_DIR / 'location'
MINE_PRJ_FILE = DATA_DIR / 'shp' / '锦界矿边界.prj'
DEFAULT_SOURCE_CRS = 'EPSG:2421'
DEFAULT_WORKFACE_NAME = '本地钻孔数据'


@dataclass
class BoreholeLocation:
    name: str
    x: float
    y: float
    z: float


def normalize_name(name: str) -> str:
    """标准化钻孔名称以提升跨表匹配容错。"""
    return ''.join(char for char in str(name).strip().upper() if char.isalnum())


def read_cell_text(value: Any) -> str:
    """将 Excel 单元格内容转换为清理后的文本。"""
    if value is None:
        return ''
    return str(value).strip()


def read_float(value: Any, default: float = 0) -> float:
    """安全读取 Excel 数值单元格。"""
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def resolve_column(header: list[str], aliases: tuple[str, ...], required: bool = True) -> int | None:
    """按字段别名解析列索引。"""
    normalized = {name.strip().lower(): index for index, name in enumerate(header) if name}
    for alias in aliases:
        index = normalized.get(alias.strip().lower())
        if index is not None:
            return index
    if required:
        raise ValueError(f'Excel 缺少列: {aliases}')
    return None


def discover_layer_files() -> list[Path]:
    """扫描本地钻孔分层 Excel 文件。"""
    if not BOREHOLE_DIR.exists():
        return []
    return sorted(path for path in BOREHOLE_DIR.glob('*.xlsx') if not path.name.startswith('~$'))


def discover_location_file() -> Path | None:
    """扫描本地钻孔位置 Excel 文件。"""
    if not LOCATION_DIR.exists():
        return None
    return next((path for path in sorted(LOCATION_DIR.glob('*.xlsx')) if not path.name.startswith('~$')), None)


def build_transformer() -> Transformer:
    """根据矿区 PRJ 或默认坐标系构造 WGS84 转换器。"""
    source_crs = CRS.from_string(DEFAULT_SOURCE_CRS)
    if MINE_PRJ_FILE.exists():
        prj_text = MINE_PRJ_FILE.read_text(encoding='utf-8', errors='ignore').strip()
        if prj_text:
            source_crs = CRS.from_wkt(prj_text)
    return Transformer.from_crs(source_crs, CRS.from_epsg(4326), always_xy=True)


def is_lonlat(lon: float, lat: float) -> bool:
    """判断坐标是否已经是经纬度。"""
    return -180 <= lon <= 180 and -90 <= lat <= 90


def convert_point_to_wgs84(x: float, y: float, transformer: Transformer) -> tuple[float, float] | None:
    """将钻孔平面坐标转换为 WGS84 坐标。"""
    if is_lonlat(x, y):
        return x, y
    lon, lat = transformer.transform(x, y)
    if is_lonlat(lon, lat):
        return lon, lat
    lon_swapped, lat_swapped = transformer.transform(y, x)
    if is_lonlat(lon_swapped, lat_swapped):
        return lon_swapped, lat_swapped
    return None


def load_location_index() -> dict[str, BoreholeLocation]:
    """读取钻孔位置表并返回按名称索引的位置字典。"""
    location_file = discover_location_file()
    if not location_file:
        return {}

    workbook = openpyxl.load_workbook(location_file, read_only=True, data_only=True)
    sheet = workbook.active
    header = [read_cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_name = resolve_column(header, LOCATION_FIELD_ALIASES['name'])
    index_x = resolve_column(header, LOCATION_FIELD_ALIASES['x'])
    index_y = resolve_column(header, LOCATION_FIELD_ALIASES['y'])
    index_z = resolve_column(header, LOCATION_FIELD_ALIASES['z'], required=False)

    result: dict[str, BoreholeLocation] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = read_cell_text(row[index_name])
        if not name:
            continue
        result[name] = BoreholeLocation(
            name=name,
            x=read_float(row[index_x]),
            y=read_float(row[index_y]),
            z=read_float(row[index_z]) if index_z is not None else 0,
        )
    workbook.close()
    return result


def match_location(name: str, locations: dict[str, BoreholeLocation]) -> BoreholeLocation | None:
    """按原始名称和标准化名称匹配钻孔位置。"""
    if name in locations:
        return locations[name]
    normalized = normalize_name(name)
    for location_name, location in locations.items():
        if normalize_name(location_name) == normalized:
            return location
    return None


def pick_layer_color(layer_name: str) -> str:
    """按地层名称给出分层默认颜色。"""
    for keyword, color in LAYER_COLORS.items():
        if keyword in layer_name:
            return color
    return '#23d18b'


def parse_layer_file(layer_file: Path) -> dict[str, list[dict[str, Any]]]:
    """解析单个钻孔分层 Excel 文件。"""
    workbook = openpyxl.load_workbook(layer_file, read_only=True, data_only=True)
    sheet = workbook.active
    header = [read_cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_name = resolve_column(header, LAYER_FIELD_ALIASES['borehole_name'])
    index_layer = resolve_column(header, LAYER_FIELD_ALIASES['layer_name'])
    index_depth = resolve_column(header, LAYER_FIELD_ALIASES['depth'])
    index_thickness = resolve_column(header, LAYER_FIELD_ALIASES['thickness'])

    result: dict[str, list[dict[str, Any]]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        borehole_name = read_cell_text(row[index_name])
        layer_name = read_cell_text(row[index_layer])
        if not borehole_name or not layer_name:
            continue
        bottom_depth = read_float(row[index_depth])
        thickness = read_float(row[index_thickness])
        layer = {
            'layer_name': layer_name,
            'top_depth': round(max(bottom_depth - thickness, 0), 4),
            'thickness': round(thickness, 4),
            'bottom_depth': round(bottom_depth, 4),
            'color': pick_layer_color(layer_name),
        }
        result.setdefault(borehole_name, []).append(layer)
    workbook.close()
    return result


def load_layer_groups() -> dict[str, list[dict[str, Any]]]:
    """加载并合并全部钻孔分层文件。"""
    merged: dict[str, list[dict[str, Any]]] = {}
    for layer_file in discover_layer_files():
        parsed = parse_layer_file(layer_file)
        for borehole_name, layers in parsed.items():
            merged.setdefault(borehole_name, []).extend(layers)
    return merged


def build_borehole_item(name: str, layers: list[dict[str, Any]], location: BoreholeLocation | None, transformer: Transformer) -> dict[str, Any]:
    """组装入库用钻孔对象。"""
    sorted_layers = sorted(layers, key=lambda item: (item['top_depth'], item['bottom_depth']))
    for sort_order, layer in enumerate(sorted_layers, start=1):
        layer['sort_order'] = sort_order

    lon = 0.0
    lat = 0.0
    elevation = 0.0
    if location:
        converted = convert_point_to_wgs84(location.x, location.y, transformer)
        if converted:
            lon, lat = converted
            elevation = location.z

    return {
        'borehole_code': name,
        'name': name,
        'longitude': round(lon, 8),
        'latitude': round(lat, 8),
        'elevation': round(elevation, 3),
        'depth_total': max((layer['bottom_depth'] for layer in sorted_layers), default=0),
        'workface_name': DEFAULT_WORKFACE_NAME,
        'remark': '由 backend/data/boreholes 与 backend/data/location 导入 PostGIS',
        'layers': sorted_layers,
    }


def resolve_workface_name(geom: Point | None, fallback: str = DEFAULT_WORKFACE_NAME) -> str:
    """按钻孔空间位置匹配工作面名称。"""
    if geom is None:
        return fallback

    from apps.boundaries.models import BoundaryRegion

    matched_name = (
        BoundaryRegion.objects
        .filter(type=BoundaryRegion.TYPE_WORKFACE, geom__contains=geom)
        .order_by('name')
        .values_list('name', flat=True)
        .first()
    )
    if matched_name:
        return matched_name

    nearest_name = (
        BoundaryRegion.objects
        .filter(type=BoundaryRegion.TYPE_WORKFACE, geom__isnull=False)
        .annotate(distance=Distance('geom', geom))
        .order_by('distance', 'name')
        .values_list('name', flat=True)
        .first()
    )
    return nearest_name or fallback


def parse_boreholes_from_files() -> list[dict[str, Any]]:
    """解析本地 Excel 数据为钻孔对象列表。"""
    transformer = build_transformer()
    locations = load_location_index()
    layer_groups = load_layer_groups()
    result = []
    for name, layers in sorted(layer_groups.items()):
        location = match_location(name, locations)
        result.append(build_borehole_item(name, layers, location, transformer))
    return result


@transaction.atomic
def import_boreholes_from_excel(remove_missing: bool = True) -> dict[str, int]:
    """将本地 Excel 钻孔数据导入数据库。"""
    parsed = parse_boreholes_from_files()
    seen_codes: set[str] = set()
    total_layers = 0

    for item in parsed:
        longitude = item['longitude']
        latitude = item['latitude']
        geom = Point(longitude, latitude, srid=4326) if not (longitude == 0 and latitude == 0) else None
        workface_name = resolve_workface_name(geom, item['workface_name'])
        borehole, _ = Borehole.objects.update_or_create(
            borehole_code=item['borehole_code'],
            defaults={
                'name': item['name'],
                'longitude': longitude,
                'latitude': latitude,
                'elevation': item['elevation'],
                'depth_total': item['depth_total'],
                'workface_name': workface_name,
                'remark': item['remark'],
                'geom': geom,
            },
        )

        borehole.layers.all().delete()
        layers = [
            BoreholeLayer(
                borehole=borehole,
                layer_name=layer['layer_name'],
                top_depth=layer['top_depth'],
                thickness=layer['thickness'],
                bottom_depth=layer['bottom_depth'],
                color=layer['color'],
                sort_order=layer['sort_order'],
            )
            for layer in item['layers']
        ]
        BoreholeLayer.objects.bulk_create(layers)
        total_layers += len(layers)
        seen_codes.add(item['borehole_code'])

    if remove_missing:
        Borehole.objects.exclude(borehole_code__in=seen_codes).delete()

    return {'boreholes': len(parsed), 'layers': total_layers}


@transaction.atomic
def sync_borehole_workface_names() -> dict[str, int]:
    """根据工作面边界回填钻孔 workface_name。"""
    changed: list[Borehole] = []
    total = 0

    for borehole in Borehole.objects.exclude(geom__isnull=True).all():
        total += 1
        fallback_name = borehole.workface_name or DEFAULT_WORKFACE_NAME
        matched_name = resolve_workface_name(borehole.geom, fallback_name)
        if matched_name != borehole.workface_name:
            borehole.workface_name = matched_name
            changed.append(borehole)

    if changed:
        Borehole.objects.bulk_update(changed, ['workface_name'])

    return {'checked': total, 'updated': len(changed)}


def serialize_queryset(queryset) -> list[dict[str, Any]]:
    """序列化钻孔查询集。"""
    return BoreholeSerializer(queryset, many=True).data


def load_boreholes() -> list[dict[str, Any]]:
    """读取数据库中的全部钻孔数据。"""
    queryset = Borehole.objects.prefetch_related('layers').all()
    return serialize_queryset(queryset)


def get_borehole_list(keyword: str | None = None, workface: str | None = None) -> list[dict[str, Any]]:
    """按条件返回钻孔列表。"""
    queryset = Borehole.objects.prefetch_related('layers').all()
    if keyword:
        queryset = queryset.filter(name__icontains=keyword)
    if workface:
        queryset = queryset.filter(workface_name=workface)
    return serialize_queryset(queryset)


def get_borehole_detail(borehole_id: str) -> dict[str, Any] | None:
    """按编号返回单个钻孔详情。"""
    borehole = Borehole.objects.prefetch_related('layers').filter(id=borehole_id).first()
    if not borehole:
        return None
    return BoreholeSerializer(borehole).data


def get_borehole_layers(borehole_id: str) -> list[dict[str, Any]]:
    """按编号返回单个钻孔分层。"""
    detail = get_borehole_detail(borehole_id)
    return detail['layers'] if detail else []


def get_borehole_geojson() -> dict[str, Any]:
    """返回钻孔 GeoJSON 点集。"""
    features = []
    for item in Borehole.objects.all():
        if item.longitude == 0 and item.latitude == 0:
            continue
        features.append({
            'type': 'Feature',
            'geometry': {'type': 'Point', 'coordinates': [item.longitude, item.latitude, item.elevation]},
            'properties': {
                'id': str(item.id),
                'borehole_code': item.borehole_code,
                'name': item.name,
                'depth_total': item.depth_total,
                'workface_name': item.workface_name,
            },
        })
    return {'type': 'FeatureCollection', 'features': features}
