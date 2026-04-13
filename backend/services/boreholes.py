from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

import openpyxl
from pyproj import CRS, Transformer

from config import DATA_DIR

LAYER_FIELD_ALIASES = {
    'borehole_name': ('钻孔名称', '孔号', 'borehole_name', 'name'),
    'layer_name': ('地层名称', '层位', 'layer_name'),
    'depth': ('深度', 'bottom_depth', 'depth'),
    'thickness': ('厚度', 'thickness'),
}

LOCATION_FIELD_ALIASES = {
    'name': ('name', '钻孔名称', '孔号', 'borehole_name'),
    'x': ('x', 'X', '经度', '东坐标', 'lon', 'longitude'),
    'y': ('y', 'Y', '纬度', '北坐标', 'lat', 'latitude'),
    'z': ('z', 'Z', '高程', 'elevation'),
    'workface_name': ('工作面', '工作面名称', 'workface', 'workface_name'),
}

LAYER_COLORS = {
    '煤': '#f2c94c',
    '砂': '#56ccf2',
    '泥': '#bb6bd9',
    '土': '#f2994a',
    '风积': '#6fcf97',
}

BOREHOLE_DIR = DATA_DIR / 'boreholes'
LOCATION_DIR = DATA_DIR / 'location'
MINE_PRJ_FILE = DATA_DIR / 'shp' / '锦界矿边界.prj'
DEFAULT_SOURCE_CRS = 'EPSG:2421'
DEFAULT_WORKFACE_NAME = '本地钻孔数据'


@dataclass
class BoreholeLocation:
    name: str
    x: float
    y: float
    z: float
    workface_name: str


def normalize_name(name: str) -> str:
    return ''.join(char for char in str(name).strip().upper() if char.isalnum())


def read_cell_text(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def read_float(value: Any, default: float = 0) -> float:
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def resolve_column(header: list[str], aliases: tuple[str, ...], required: bool = True) -> int | None:
    normalized = {name.strip().lower(): index for index, name in enumerate(header) if name}
    for alias in aliases:
        index = normalized.get(alias.strip().lower())
        if index is not None:
            return index
    if required:
        raise ValueError(f'Excel missing column: {aliases}')
    return None


def discover_layer_files() -> list[Path]:
    if not BOREHOLE_DIR.exists():
        return []
    return sorted(path for path in BOREHOLE_DIR.glob('*.xlsx') if not path.name.startswith('~$'))


def discover_location_file() -> Path | None:
    if not LOCATION_DIR.exists():
        return None
    return next((path for path in sorted(LOCATION_DIR.glob('*.xlsx')) if not path.name.startswith('~$')), None)


def build_transformer() -> Transformer:
    source_crs = CRS.from_string(DEFAULT_SOURCE_CRS)
    if MINE_PRJ_FILE.exists():
        prj_text = MINE_PRJ_FILE.read_text(encoding='utf-8', errors='ignore').strip()
        if prj_text:
            source_crs = CRS.from_wkt(prj_text)
    return Transformer.from_crs(source_crs, CRS.from_epsg(4326), always_xy=True)


def is_lonlat(lon: float, lat: float) -> bool:
    return -180 <= lon <= 180 and -90 <= lat <= 90


def convert_point_to_wgs84(x: float, y: float, transformer: Transformer) -> tuple[float, float] | None:
    if is_lonlat(x, y):
        return x, y
    lon, lat = transformer.transform(x, y)
    if is_lonlat(lon, lat):
        return lon, lat
    lon_swapped, lat_swapped = transformer.transform(y, x)
    if is_lonlat(lon_swapped, lat_swapped):
        return lon_swapped, lat_swapped
    return None


def load_location_index() -> dict[str, BoreholeLocation]:
    location_file = discover_location_file()
    if not location_file:
        return {}

    workbook = openpyxl.load_workbook(location_file, read_only=True, data_only=True)
    sheet = workbook.active
    header = [read_cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_name = resolve_column(header, LOCATION_FIELD_ALIASES['name'])
    index_x = resolve_column(header, LOCATION_FIELD_ALIASES['x'])
    index_y = resolve_column(header, LOCATION_FIELD_ALIASES['y'])
    index_z = resolve_column(header, LOCATION_FIELD_ALIASES['z'], required=False)
    index_workface_name = resolve_column(header, LOCATION_FIELD_ALIASES['workface_name'], required=False)

    result: dict[str, BoreholeLocation] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = read_cell_text(row[index_name])
        if not name:
            continue
        workface_name = read_cell_text(row[index_workface_name]) if index_workface_name is not None else ''
        result[name] = BoreholeLocation(
            name=name,
            x=read_float(row[index_x]),
            y=read_float(row[index_y]),
            z=read_float(row[index_z]) if index_z is not None else 0,
            workface_name=workface_name,
        )
    workbook.close()
    return result


def match_location(name: str, locations: dict[str, BoreholeLocation]) -> BoreholeLocation | None:
    if name in locations:
        return locations[name]
    normalized = normalize_name(name)
    for location_name, location in locations.items():
        if normalize_name(location_name) == normalized:
            return location
    return None


def pick_layer_color(layer_name: str) -> str:
    for keyword, color in LAYER_COLORS.items():
        if keyword in layer_name:
            return color
    return '#23d18b'


def parse_layer_file(layer_file: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(layer_file, read_only=True, data_only=True)
    sheet = workbook.active
    header = [read_cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_name = resolve_column(header, LAYER_FIELD_ALIASES['borehole_name'])
    index_layer = resolve_column(header, LAYER_FIELD_ALIASES['layer_name'])
    index_depth = resolve_column(header, LAYER_FIELD_ALIASES['depth'])
    index_thickness = resolve_column(header, LAYER_FIELD_ALIASES['thickness'])

    result: dict[str, list[dict[str, Any]]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        borehole_name = read_cell_text(row[index_name])
        layer_name = read_cell_text(row[index_layer])
        if not borehole_name or not layer_name:
            continue
        bottom_depth = read_float(row[index_depth])
        thickness = read_float(row[index_thickness])
        layer = {
            'layer_name': layer_name,
            'top_depth': round(max(bottom_depth - thickness, 0), 4),
            'thickness': round(thickness, 4),
            'bottom_depth': round(bottom_depth, 4),
            'color': pick_layer_color(layer_name),
        }
        result.setdefault(borehole_name, []).append(layer)
    workbook.close()
    return result


def load_layer_groups() -> dict[str, list[dict[str, Any]]]:
    merged: dict[str, list[dict[str, Any]]] = {}
    for layer_file in discover_layer_files():
        parsed = parse_layer_file(layer_file)
        for borehole_name, layers in parsed.items():
            merged.setdefault(borehole_name, []).extend(layers)
    return merged


def build_borehole_id(name: str) -> str:
    normalized = normalize_name(name) or name
    return str(uuid5(NAMESPACE_URL, f'minescope3d:borehole:{normalized}'))


def build_layer_id(borehole_id: str, sort_order: int, layer_name: str) -> str:
    return str(uuid5(NAMESPACE_URL, f'minescope3d:borehole-layer:{borehole_id}:{sort_order}:{layer_name}'))


def build_borehole_item(
    name: str,
    layers: list[dict[str, Any]],
    location: BoreholeLocation | None,
    transformer: Transformer,
) -> dict[str, Any]:
    sorted_layers = sorted(layers, key=lambda item: (item['top_depth'], item['bottom_depth']))

    borehole_id = build_borehole_id(name)
    for sort_order, layer in enumerate(sorted_layers, start=1):
        layer['id'] = build_layer_id(borehole_id, sort_order, layer['layer_name'])
        layer['sort_order'] = sort_order

    lon = 0.0
    lat = 0.0
    elevation = 0.0
    workface_name = DEFAULT_WORKFACE_NAME
    if location:
        converted = convert_point_to_wgs84(location.x, location.y, transformer)
        if converted:
            lon, lat = converted
            elevation = location.z
        if location.workface_name:
            workface_name = location.workface_name

    return {
        'id': borehole_id,
        'borehole_code': name,
        'name': name,
        'longitude': round(lon, 8),
        'latitude': round(lat, 8),
        'elevation': round(elevation, 3),
        'depth_total': max((layer['bottom_depth'] for layer in sorted_layers), default=0),
        'workface_name': workface_name,
        'remark': 'Loaded from backend/data/boreholes and backend/data/location files',
        'layers': sorted_layers,
    }


def parse_boreholes_from_files() -> list[dict[str, Any]]:
    transformer = build_transformer()
    locations = load_location_index()
    layer_groups = load_layer_groups()

    result = []
    for name, layers in sorted(layer_groups.items()):
        location = match_location(name, locations)
        result.append(build_borehole_item(name, layers, location, transformer))
    return result


def import_boreholes_from_excel(remove_missing: bool = True) -> dict[str, int]:
    del remove_missing
    parsed = parse_boreholes_from_files()
    total_layers = sum(len(item.get('layers', [])) for item in parsed)
    return {'boreholes': len(parsed), 'layers': total_layers}


def sync_borehole_workface_names() -> dict[str, int]:
    parsed = parse_boreholes_from_files()
    return {'checked': len(parsed), 'updated': 0}


def load_boreholes() -> list[dict[str, Any]]:
    return parse_boreholes_from_files()


def get_borehole_list(keyword: str | None = None, workface: str | None = None) -> list[dict[str, Any]]:
    rows = load_boreholes()
    if keyword:
        lower_keyword = keyword.lower()
        rows = [item for item in rows if lower_keyword in item['name'].lower()]
    if workface:
        rows = [item for item in rows if item.get('workface_name') == workface]
    return rows


def get_borehole_detail(borehole_id: str) -> dict[str, Any] | None:
    for item in load_boreholes():
        if item['id'] == borehole_id or item['borehole_code'] == borehole_id:
            return item
    return None


def get_borehole_layers(borehole_id: str) -> list[dict[str, Any]]:
    detail = get_borehole_detail(borehole_id)
    return detail['layers'] if detail else []


def get_borehole_geojson() -> dict[str, Any]:
    features = []
    for item in load_boreholes():
        if item['longitude'] == 0 and item['latitude'] == 0:
            continue
        features.append(
            {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': [item['longitude'], item['latitude'], item['elevation']],
                },
                'properties': {
                    'id': item['id'],
                    'borehole_code': item['borehole_code'],
                    'name': item['name'],
                    'depth_total': item['depth_total'],
                    'workface_name': item['workface_name'],
                },
            }
        )
    return {'type': 'FeatureCollection', 'features': features}
